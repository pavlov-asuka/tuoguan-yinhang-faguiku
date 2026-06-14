from __future__ import annotations

import csv
import hashlib
import html
import io
import json
import re
import shutil
import ssl
import tempfile
import time
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from lxml import etree
from lxml import html as lxml_html
from openpyxl import load_workbook
from pypdf import PdfReader

from build_search_index import build_index


ROOT = Path(__file__).resolve().parents[1]
CATALOG_PATH = ROOT / "中国商业银行托管业务条线大法规库总目录.md"
RAW_ROOT = ROOT / "01_法规原文库"
TEXT_ROOT = ROOT / "02_文本抽取库"
META_ROOT = ROOT / "03_元数据台账"
SOURCE_OVERRIDES_PATH = META_ROOT / "source_overrides.json"
ENTRY_ROOT = ROOT / "00_入口与索引"
UNRESOLVED_PATH = ENTRY_ROOT / "unresolved.md"
BUILD_CACHE_ROOT: Path | None = None
UPDATE_LOG_PATH = ENTRY_ROOT / "更新记录.md"
MANUAL_UPDATE_HEADING = "## 人工维护记录"
AUTO_UPDATE_HEADING = "## 自动重建记录"

DEFAULT_MANUAL_UPDATE_LOG = """- 2026-06-14：重做 `00_入口与索引/总目录.html` 为法规库工作台，提供目录树、筛选表格、详情面板和相对路径原文链接；同步明确“总收录记录”“有本地原文文件”“官方来源覆盖”等指标口径。
- 2026-06-14：将 TB181“业绩比较基准相关指标解释、基准收益率计算参考及基准示例”纳入正式规则，保留一类库、二类库和《公募基金业绩比较基准要素库运作说明》本地原文与文本。
- 2026-06-14：完成 unresolved 清理和大目录完整性复核；官方发布资料不再使用“辅助资料”低效力分类，按正式规则、重要参考规则、规则组索引或官方入口处理。
- 2026-06-14：确认 TB101/TB184 等手册类资料作为重要参考规则保留，不降级为资料入口。
- 2026-06-13：补强本地检索索引，重要参考规则、规则组索引和官方入口等无正文记录可通过元数据块召回；同步修复 Windows 控制台输出兼容问题。
- 2026-06-12：按“中国商业银行托管业务条线大法规库”口径重构法规库，归档旧公募基金托管库并复用已核验原文和文本。"""

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)

INTERNAL_KEYWORDS = (
    "内部落库位",
    "内部模板",
    "内部制度",
    "样表",
    "流程图",
    "知识运营",
    "制度模板",
)

COMPOSITE_TITLE_KEYWORDS = (
    "相关",
    "配套",
    "规则组",
    "清单",
    "入口",
    "案例",
    "模板",
    "样表",
    "手册",
    "体系",
    "指引库",
    "报告规则",
    "职责规则",
    "税务规则",
    "差异对照",
    "总目录",
)

TAG_PRODUCTS = (
    "公募基金",
    "私募基金",
    "证券期货资管",
    "银行理财",
    "信托",
    "保险资金",
    "保险资管",
    "养老金",
    "社保",
    "年金",
    "QDII",
    "QFII",
    "REITs",
    "衍生品",
    "客户资金",
)

TAG_LINES = (
    "准入",
    "前台",
    "账户保管",
    "账户",
    "保管",
    "核算估值",
    "估值",
    "核算",
    "清算交收",
    "清算",
    "交收",
    "投资监督",
    "披露报送",
    "披露",
    "报送",
    "IT数据",
    "数据",
    "合同法务",
    "合同",
    "法务",
    "风控内审",
    "风控",
    "内控",
    "审计",
    "税务",
    "反洗钱",
    "销售",
)

TAG_MARKETS = (
    "交易所",
    "银行间",
    "场外",
    "跨境",
    "港股通",
    "债券通",
    "外汇",
    "期货",
    "衍生品",
)

OFFICIAL_ENTRY_URLS = [
    (("国家法律法规数据库",), ["https://flk.npc.gov.cn/"]),
    (("中国政府网",), ["https://www.gov.cn/zhengce/"]),
    (("中国人民银行",), ["https://www.pbc.gov.cn/"]),
    (("国家金融监督管理总局",), ["https://www.nfra.gov.cn/"]),
    (("中国证监会", "证监会"), ["https://www.csrc.gov.cn/"]),
    (("国家外汇管理局", "外汇局"), ["https://www.safe.gov.cn/"]),
    (("财政部",), ["https://www.mof.gov.cn/"]),
    (("国家税务总局", "税务总局"), ["https://www.chinatax.gov.cn/"]),
    (("中国证券投资基金业协会", "中基协", "AMBERS"), ["https://www.amac.org.cn/"]),
    (("中国银行业协会", "中银协"), ["https://www.china-cba.net/"]),
    (("中国证券业协会", "中证协"), ["https://www.sac.net.cn/"]),
    (("中国期货业协会", "期货业协会"), ["https://www.cfachina.org/"]),
    (("中国保险资产管理业协会", "保险资管业协会"), ["https://www.iamac.org.cn/"]),
    (("中国银行间市场交易商协会", "交易商协会"), ["https://www.nafmii.org.cn/"]),
    (("上海证券交易所",), ["https://www.sse.com.cn/"]),
    (("深圳证券交易所",), ["https://www.szse.cn/"]),
    (("北京证券交易所",), ["https://www.bse.cn/"]),
    (("港交所", "香港结算"), ["https://www.hkex.com.hk/"]),
    (("中国结算", "中登", "中国证券登记结算"), ["https://www.chinaclear.cn/"]),
    (("上海清算所", "银行间市场清算所"), ["https://www.shclearing.com.cn/"]),
    (("中央结算", "中债", "中央国债登记结算"), ["https://www.chinabond.com.cn/"]),
    (("外汇交易中心",), ["https://www.chinamoney.com.cn/"]),
    (("中国证券投资基金电子披露网站", "电子披露网站"), ["https://eid.csrc.gov.cn/fund/"]),
    (("资产托管网上服务平台",), ["https://www.chinaclear.cn/"]),
    (("个人养老金基金行业平台", "FIRM"), ["https://www.chinaclear.cn/", "https://www.amac.org.cn/"]),
    (("FISP",), ["https://www.amac.org.cn/"]),
    (("交易所", "交易所纪律处分"), ["https://www.sse.com.cn/", "https://www.szse.cn/", "https://www.bse.cn/"]),
    (("地方证监局",), ["https://www.csrc.gov.cn/"]),
    (("资管新规",), ["https://www.gov.cn/zhengce/", "https://www.pbc.gov.cn/"]),
    (("商业银行资产托管业务指引",), ["https://www.china-cba.net/"]),
    (("基金估值",), ["https://www.amac.org.cn/"]),
    (("会计准则",), ["https://kjs.mof.gov.cn/", "https://fgk.mof.gov.cn/"]),
    (("证券期货",), ["https://www.csrc.gov.cn/"]),
    (("私募资产管理", "私募基金"), ["https://www.csrc.gov.cn/", "https://www.amac.org.cn/"]),
    (("银行理财", "理财公司"), ["https://www.nfra.gov.cn/"]),
    (("信托",), ["https://www.nfra.gov.cn/"]),
    (("保险资金", "保险资管", "保险资产管理"), ["https://www.nfra.gov.cn/", "https://www.iamac.org.cn/"]),
    (("企业年金", "职业年金", "社保基金", "养老金"), ["https://www.mohrss.gov.cn/", "https://www.ssf.gov.cn/"]),
    (("QDII", "QFII", "外汇管理"), ["https://www.safe.gov.cn/", "https://www.csrc.gov.cn/"]),
    (("沪港通", "深港通", "港股通", "互联互通"), ["https://www.sse.com.cn/", "https://www.szse.cn/", "https://www.chinaclear.cn/", "https://www.hkex.com.hk/"]),
    (("资产证券化", "资产支持证券", "ABS", "ABN", "资产支持票据"), ["https://www.csrc.gov.cn/", "https://www.sse.com.cn/", "https://www.szse.cn/", "https://www.nafmii.org.cn/"]),
    (("转融通", "融资融券"), ["https://www.csrc.gov.cn/", "https://www.sse.com.cn/", "https://www.szse.cn/", "https://www.chinaclear.cn/"]),
    (("银行间债券", "债券登记托管", "债券借贷", "质押式回购", "买断式回购"), ["https://www.chinabond.com.cn/", "https://www.chinamoney.com.cn/", "https://www.shclearing.com.cn/"]),
    (("人民币银行结算账户", "大额支付系统", "小额支付系统", "网上支付跨行清算", "人民币跨境支付", "反洗钱处罚", "支付结算处罚", "征信处罚"), ["https://www.pbc.gov.cn/"]),
    (("客户尽职调查", "客户身份资料", "交易记录保存"), ["https://www.pbc.gov.cn/", "https://www.nfra.gov.cn/", "https://www.csrc.gov.cn/"]),
    (("基金合同", "基金招募说明书", "产品资料概要", "公募基金定期报告", "货币市场基金信息披露", "公募基金投资信用衍生品"), ["https://www.csrc.gov.cn/", "https://www.amac.org.cn/"]),
    (("托管协议", "运营外包协议", "资金监管协议", "账户服务协议", "电子合同服务协议", "三方/四方业务协议", "境外次托管协议"), ["https://www.csrc.gov.cn/", "https://www.nfra.gov.cn/", "https://www.amac.org.cn/", "https://www.chinaclear.cn/", "https://www.safe.gov.cn/"]),
    (("产品设立", "产品注册", "产品备案", "发行", "募集", "产品运作", "终止清算"), ["https://www.csrc.gov.cn/", "https://www.amac.org.cn/", "https://www.nfra.gov.cn/", "https://www.safe.gov.cn/"]),
    (("业绩比较基准", "基准收益率", "基准示例"), ["https://www.csrc.gov.cn/", "https://www.amac.org.cn/", "https://www.csindex.com.cn/"]),
    (("SJSMX1", "BJSMX1", "ETF申赎", "业务确认接口", "接口规范"), ["https://www.chinaclear.cn/", "https://www.sse.com.cn/", "https://www.szse.cn/", "https://www.bse.cn/", "https://www.csindex.com.cn/"]),
    (("现金管理类理财", "银行业金融机构", "商业银行合规风险", "全面风险管理", "业务连续性监管", "信息科技风险"), ["https://www.nfra.gov.cn/"]),
    (("估值差错", "迟算漏算", "错划款", "异常交收", "监督遗漏", "披露错误", "报送错误", "合同瑕疵", "权限滥用", "外包风险"), ["https://www.csrc.gov.cn/", "https://www.nfra.gov.cn/", "https://www.amac.org.cn/", "https://www.china-cba.net/", "https://www.chinaclear.cn/"]),
    (("全国银行间同业拆借中心", "同业拆借中心"), ["https://www.chinamoney.com.cn/"]),
    (("中证指数",), ["https://www.csindex.com.cn/"]),
    (("期货保证金监控中心", "期货账户", "期货结算"), ["https://www.cfmmc.com/"]),
    (("融资融券", "转融通", "中证金融"), ["https://www.csf.com.cn/", "https://www.sse.com.cn/", "https://www.szse.cn/"]),
]

TITLE_ALIASES = {
    "证券投资基金法": "中华人民共和国证券投资基金法",
    "证券法": "中华人民共和国证券法",
    "商业银行法": "中华人民共和国商业银行法",
    "银行业监督管理法": "中华人民共和国银行业监督管理法",
    "信托法": "中华人民共和国信托法",
    "期货和衍生品法": "中华人民共和国期货和衍生品法",
    "中国人民银行法": "中华人民共和国中国人民银行法",
    "保险法": "中华人民共和国保险法",
    "公司法": "中华人民共和国公司法",
    "合伙企业法": "中华人民共和国合伙企业法",
    "企业破产法": "中华人民共和国企业破产法",
    "民法典": "中华人民共和国民法典",
    "电子签名法": "中华人民共和国电子签名法",
    "仲裁法": "中华人民共和国仲裁法",
    "民事诉讼法": "中华人民共和国民事诉讼法",
    "反洗钱法": "中华人民共和国反洗钱法",
    "网络安全法": "中华人民共和国网络安全法",
    "数据安全法": "中华人民共和国数据安全法",
    "个人信息保护法": "中华人民共和国个人信息保护法",
    "消费者权益保护法": "中华人民共和国消费者权益保护法",
    "密码法": "中华人民共和国密码法",
    "国家安全法": "中华人民共和国国家安全法",
    "反恐怖主义法": "中华人民共和国反恐怖主义法",
    "增值税法": "中华人民共和国增值税法",
    "企业所得税法": "中华人民共和国企业所得税法",
    "个人所得税法": "中华人民共和国个人所得税法",
    "印花税法": "中华人民共和国印花税法",
    "税收征收管理法": "中华人民共和国税收征收管理法",
}

ISSUER_BY_HINT = {
    "全国人大": "全国人民代表大会常务委员会",
    "法律": "全国人民代表大会常务委员会",
    "行政法规": "国务院",
    "司法解释": "最高人民法院",
    "证监会": "中国证监会",
    "人民银行": "中国人民银行",
    "外汇局": "国家外汇管理局",
    "财政部": "财政部",
    "税务总局": "国家税务总局",
    "交易所": "证券交易所",
    "中登": "中国证券登记结算有限责任公司",
    "基础设施": "市场基础设施",
    "自律规则": "行业自律组织",
    "自律入口": "行业自律组织",
}

FIELDS = [
    "id",
    "legacy_id",
    "priority",
    "category",
    "title",
    "layer",
    "doc_type",
    "issuer",
    "rule_no",
    "publish_date",
    "effective_date",
    "current_status",
    "record_role",
    "ingest_status",
    "is_core",
    "source_type",
    "product_tags",
    "business_line_tags",
    "market_tags",
    "business_tags",
    "catalog_paths",
    "key_obligations",
    "source_url",
    "local_path",
    "text_path",
    "file_type",
    "downloaded_count",
    "notes",
    "errors",
]


@dataclass
class CatalogItem:
    title: str
    priority: str
    doc_type: str
    body: str
    category: str
    catalog_path: str
    top_code: str
    line_no: int


@dataclass
class RuleRecord:
    key: str
    id: str = ""
    legacy_id: str = ""
    title: str = ""
    priority: str = "P2"
    category: str = ""
    layer: str = ""
    doc_type: str = ""
    issuer: str = ""
    rule_no: str = ""
    publish_date: str = ""
    effective_date: str = ""
    current_status: str = "待扩展"
    record_role: str = ""
    ingest_status: str = ""
    is_core: str = "否"
    source_type: str = ""
    product_tags: list[str] = field(default_factory=list)
    business_line_tags: list[str] = field(default_factory=list)
    market_tags: list[str] = field(default_factory=list)
    business_tags: list[str] = field(default_factory=list)
    catalog_paths: list[str] = field(default_factory=list)
    key_obligations: str = ""
    source_url: str = ""
    local_path: str = ""
    text_path: str = ""
    file_type: str = ""
    downloaded_count: int = 0
    notes: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    legacy_row: dict[str, Any] | None = None
    first_line: int = 0


BINDING_SINGLE_RULE_IDS = {
    "TB037",
    "TB080",
    "TB101",
    "TB114",
    "TB115",
    "TB117",
    "TB118",
    "TB119",
    "TB120",
    "TB121",
    "TB126",
    "TB165",
    "TB184",
    "TB186",
    "TB187",
    "TB189",
    "TB192",
    "TB205",
    "TB206",
    "TB221",
    "TB226",
    "TB233",
    "TB239",
    "TB265",
    "TB266",
    "TB267",
    "TB269",
    "TB270",
    "TB273",
    "TB276",
    "TB279",
    "TB283",
    "TB284",
    "TB287",
    "TB288",
    "TB310",
    "TB320",
    "TB328",
    "TB329",
}

RULE_GROUP_IDS = {
    "TB043",
    "TB058",
    "TB059",
    "TB060",
    "TB061",
    "TB062",
    "TB063",
    "TB064",
    "TB065",
    "TB075",
    "TB079",
    "TB081",
    "TB082",
    "TB085",
    "TB087",
    "TB089",
    "TB090",
    "TB102",
    "TB113",
    "TB116",
    "TB128",
    "TB150",
    "TB153",
    "TB156",
    "TB157",
    "TB158",
    "TB159",
    "TB160",
    "TB166",
    "TB170",
    "TB171",
    "TB175",
    "TB177",
    "TB178",
    "TB190",
    "TB193",
    "TB194",
    "TB195",
    "TB196",
    "TB197",
    "TB201",
    "TB202",
    "TB207",
    "TB208",
    "TB211",
    "TB213",
    "TB214",
    "TB223",
    "TB224",
    "TB225",
    "TB227",
    "TB229",
    "TB230",
    "TB234",
    "TB235",
    "TB236",
    "TB240",
    "TB241",
    "TB243",
    "TB244",
    "TB245",
    "TB246",
    "TB248",
    "TB251",
    "TB254",
    "TB255",
    "TB256",
    "TB259",
    "TB260",
    "TB261",
    "TB262",
    "TB263",
    "TB268",
    "TB271",
    "TB272",
    "TB274",
    "TB275",
    "TB277",
    "TB278",
    "TB280",
    "TB281",
    "TB282",
    "TB285",
    "TB286",
    "TB289",
    "TB290",
    "TB291",
    "TB292",
    "TB293",
    "TB295",
    "TB296",
    "TB297",
    "TB298",
    "TB299",
    "TB300",
    "TB301",
    "TB302",
    "TB303",
    "TB304",
    "TB305",
    "TB306",
    "TB307",
    "TB308",
    "TB313",
    "TB315",
    "TB317",
    "TB318",
    "TB322",
    "TB326",
    "TB327",
    "TB330",
    "TB332",
    "TB341",
    "TB344",
    "TB345",
}

MIXED_REFERENCE_RULE_IDS = {
    "TB013",
    "TB033",
    "TB034",
    "TB038",
    "TB039",
    "TB042",
    "TB047",
    "TB067",
    "TB074",
    "TB088",
    "TB091",
    "TB151",
    "TB155",
    "TB161",
    "TB167",
    "TB174",
    "TB228",
    "TB249",
    "TB250",
    "TB294",
    "TB309",
    "TB311",
    "TB312",
    "TB314",
    "TB334",
    "TB335",
    "TB339",
    "TB340",
    "TB342",
    "TB343",
    "TB346",
    "TB347",
    "TB348",
}

IMPORTANT_REFERENCE_HANDBOOK_IDS = {"TB101", "TB184"}
IMPORTANT_REFERENCE_STALE_NOTE_FRAGMENTS = (
    "辅助资料，不作为强制规则",
    "手册标明仅供参考，不作为必然依据",
)
IMPORTANT_REFERENCE_STALE_SOURCE_TYPES = {"中基协业务资料"}
FORMAL_REFERENCE_RULE_IDS = {"TB181"}
IMPORTANT_REFERENCE_RULE_IDS: set[str] = set()
IMPORTANT_REFERENCE_RULE_STALE_NOTE_FRAGMENTS = (
    "官方入口仅记入台账；不作为本地正式原文",
    "尚无本地正式原文，待月度巡检",
    "未命中规则关键词，暂按辅助资料",
    "命中重要参考规则清单",
    "重要参考规则按有效官方发布参考处理",
)
STALE_AUXILIARY_NOTE_FRAGMENTS = (
    "辅助资料",
    "规则集合不得以辅助资料状态结案",
    "不作为单项现行正式规则",
    "仅辅助资料",
)

OFFICIAL_ENTRY_RE = re.compile(r"入口|栏目|政策文件库|法规库|官方库|查询")
AUXILIARY_MATERIAL_RE = re.compile(
    r"案例|处罚|处分|监管措施|监管通报|问答|口径|申请材料|申请表|流程说明|模板|样表|培训资料|"
    r"云课堂|测试资料|报文样例|交易日历|结算日历|汇率|权益事件|停牌退市|统计|行业分析|"
    r"起草说明|答记者问|政策解读|系统说明|资料$"
)
NORMATIVE_RE = re.compile(
    r"法律|法规|法典|条例|司法解释|规则|办法|规定|指引|通知|细则|准则|规程|规范|业务指南|操作实务手册|管理|披露|报送|"
    r"登记|结算|交易|托管|清算|账户|债券|基金|资管|理财|信托|保险资金|会计|审计|税收|"
    r"税务|合规|风险|内控|职责"
)


def split_field(value: str) -> list[str]:
    return [part.strip() for part in (value or "").split(";") if part.strip()]


def join_unique(values: list[str]) -> str:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        value = value.strip()
        if value and value not in seen:
            seen.add(value)
            out.append(value)
    return "; ".join(out)


def add_note(record: RuleRecord, note: str) -> None:
    if note and note not in record.notes:
        record.notes.append(note)


def classify_record(record: RuleRecord) -> tuple[str, str, str]:
    text = f"{record.doc_type} {record.title} {record.category}"
    if record.id in BINDING_SINGLE_RULE_IDS:
        return "正式规则", "待核验原文" if not record.local_path else "已入库", "命中误分类修正清单：单项具约束力规则"
    if record.id in FORMAL_REFERENCE_RULE_IDS:
        return "正式规则", "待核验原文" if not record.local_path else "已入库", "命中正式规则白名单：官方发布正文/附件已入库"
    if record.id in IMPORTANT_REFERENCE_RULE_IDS:
        return "重要参考规则", "已核验参考", "命中重要参考规则清单：非单项强制正文但应作为业务规则参考"
    if record.id in RULE_GROUP_IDS:
        return "规则组索引", "已核验索引", "命中误分类修正清单：规则集合需拆分为具体规则"
    if record.id in MIXED_REFERENCE_RULE_IDS:
        return "重要参考规则", "已核验参考", "命中重要参考规则清单：官方发布或规则配套参考，按有效参考规则处理"
    if record.local_path and record.current_status == "现行有效":
        return "正式规则", "已入库", "已有正式原文且状态为现行有效"
    if OFFICIAL_ENTRY_RE.search(text):
        return "官方入口", "仅入口", "官方入口或栏目，不作为单项规则正文"
    if NORMATIVE_RE.search(text):
        if is_composite_title(record.title):
            if record.local_path and record.current_status == "现行有效":
                return "正式规则", "已入库", "集合型专题已引用正式原文"
            return "规则组索引", "已核验索引", "标题包含规则关键词且为集合型条目"
        return "正式规则", "待核验原文" if not record.local_path else "已入库", "标题包含规则关键词，按正式规则处理"
    if AUXILIARY_MATERIAL_RE.search(text):
        return "重要参考规则", "已核验参考", "官方发布的案例、问答、模板、培训或说明性文件按有效参考规则处理"
    return "重要参考规则", "已核验参考", "未命中正式规则关键词，按官方发布参考规则处理"


def apply_classification_guardrail(record: RuleRecord) -> None:
    if record.current_status == "历史失效":
        record.record_role = record.record_role or "正式规则"
        record.ingest_status = record.ingest_status or "已失效"
        return

    role, ingest_status, reason = classify_record(record)
    record.record_role = role
    record.ingest_status = "已入库" if record.local_path and ingest_status.startswith("待核验原文") else ingest_status
    add_note(record, f"分类口径：{reason}")

    if role == "正式规则":
        record.source_type = join_unique(
            [part for part in split_field(record.source_type) if part != "重要参考规则"]
        )
        record.notes = [
            note
            for note in record.notes
            if not any(fragment in note for fragment in IMPORTANT_REFERENCE_RULE_STALE_NOTE_FRAGMENTS)
        ]
        if record.current_status in ["辅助资料", "辅助索引", "不适用", "待扩展"] or not record.current_status:
            record.current_status = "待核验"
        if not record.local_path:
            add_note(record, "正式规则需补入官方原文并核验现行性")
    elif role == "规则组索引":
        if record.current_status in ["辅助资料", "辅助索引", "待核验", "待扩展"] or not record.current_status:
            record.current_status = "不适用"
        add_note(record, "专题索引已核验；具体适用回到已拆分正式规则正文")
    elif role == "重要参考规则":
        if record.current_status in ["辅助资料", "辅助索引", "待核验", "待扩展", "不适用"] or not record.current_status:
            record.current_status = "现行有效"
        record.source_type = join_unique(split_field(record.source_type) + ["重要参考规则"])
        record.notes = [
            note
            for note in record.notes
            if not any(fragment in note for fragment in IMPORTANT_REFERENCE_RULE_STALE_NOTE_FRAGMENTS)
        ]
        add_note(record, "重要参考规则按有效官方发布参考处理；无公开正文时按元数据和关联正式规则召回")
    elif role == "官方入口":
        if record.current_status in ["辅助资料", "辅助索引", "待核验", "待扩展"] or not record.current_status:
            record.current_status = "不适用"


def clean_notes(notes: list[str], *, auxiliary_p0: bool = False) -> list[str]:
    cleaned: list[str] = []
    replacement = "组合型P0目录已入官方入口；正式适用应回到具体规则正文核验"
    for note in notes:
        for part in split_field(note):
            if part == "官方入口已入库；不作为现行正式依据":
                continue
            if part == "P0规则尚无本地正式原文，待月度巡检":
                if auxiliary_p0 and replacement not in cleaned:
                    cleaned.append(replacement)
                continue
            if part and part not in cleaned:
                cleaned.append(part)
    return cleaned


def normalize_important_reference_handbook(record: RuleRecord) -> None:
    if record.id not in IMPORTANT_REFERENCE_HANDBOOK_IDS:
        return
    record.source_type = join_unique(
        [
            part
            for part in split_field(record.source_type)
            if part not in IMPORTANT_REFERENCE_STALE_SOURCE_TYPES
        ]
        + ["中基协重要规则参考"]
    )
    record.notes = [
        note
        for note in record.notes
        if not any(fragment in note for fragment in IMPORTANT_REFERENCE_STALE_NOTE_FRAGMENTS)
    ]


def clean_stale_auxiliary_notes(record: RuleRecord) -> None:
    record.notes = [
        note
        for note in record.notes
        if not any(fragment in note for fragment in STALE_AUXILIARY_NOTE_FRAGMENTS)
    ]


def safe_name(value: str, max_len: int = 90) -> str:
    value = html.unescape(value)
    value = re.sub(r"[<>:\"/\\|?*\r\n\t]", "_", value)
    value = re.sub(r"\s+", "", value)
    value = value.strip(" ._")
    return value[:max_len] or "untitled"


def relpath(path: Path) -> str:
    return str(path.resolve().relative_to(ROOT.resolve()))


def ensure_under(path: Path, parent: Path) -> None:
    resolved = path.resolve()
    base = parent.resolve()
    if resolved != base and not resolved.is_relative_to(base):
        raise RuntimeError(f"拒绝操作工作区之外的路径：{resolved}")


def normalize_title(value: str) -> str:
    value = TITLE_ALIASES.get(value.strip(), value.strip())
    value = re.sub(r"[《》〈〉“”\"'（）()、，,。；;：:\s·—_\-/]", "", value)
    value = value.replace("中华人民共和国", "")
    return value


def canonical_title(value: str) -> str:
    value = value.strip()
    return TITLE_ALIASES.get(value, value)


def title_candidates(value: str) -> list[str]:
    title = canonical_title(value)
    candidates = [title, value]
    if title.startswith("中华人民共和国"):
        candidates.append(title.removeprefix("中华人民共和国"))
    else:
        candidates.append("中华人民共和国" + title)
    return [normalize_title(item) for item in candidates if item]


def priority_rank(value: str) -> int:
    return {"P0": 0, "P1": 1, "P2": 2, "P3": 3}.get(value, 9)


def stronger_priority(a: str, b: str) -> str:
    return a if priority_rank(a) <= priority_rank(b) else b


def extract_title(body: str) -> str:
    quoted = re.search(r"《([^》]+)》", body)
    if quoted:
        return canonical_title(quoted.group(1).strip())
    body = re.sub(r"（[^）]*(待核验|待补齐|现行状态待核验)[^）]*）", "", body)
    body = body.strip(" ；;。")
    return canonical_title(body)


def category_from_stack(stack: dict[int, str]) -> str:
    parts: list[str] = []
    for level in [1, 2, 3]:
        if level in stack:
            parts.append(safe_name(stack[level], max_len=60))
    return "/".join(parts)


def parse_catalog() -> tuple[list[CatalogItem], dict[str, Any]]:
    lines = CATALOG_PATH.read_text(encoding="utf-8").splitlines()
    stack: dict[int, str] = {}
    top_code = ""
    items: list[CatalogItem] = []
    skipped_internal = 0

    for line_no, line in enumerate(lines, start=1):
        top = re.match(r"^# ([A-G])\. (.+)$", line)
        if top:
            top_code = top.group(1)
            stack = {1: f"{top.group(1)}_{top.group(2).strip()}"}
            continue

        heading = re.match(r"^(#{2,4})\s+(.+)$", line)
        if heading and top_code:
            level = len(heading.group(1))
            stack[level] = heading.group(2).strip()
            for stale in [k for k in stack if k > level]:
                stack.pop(stale, None)
            continue

        bullet = re.match(r"^- 【(?P<priority>P[012])｜(?P<doc_type>[^】]+)】(?P<body>.+)$", line)
        if not bullet or not top_code:
            continue

        raw = line.strip()
        second_level = stack.get(2, "")
        if second_level.startswith("G-02") or second_level.startswith("G-03"):
            skipped_internal += 1
            continue
        if any(keyword in raw for keyword in INTERNAL_KEYWORDS):
            skipped_internal += 1
            continue

        body = bullet.group("body").strip()
        title = extract_title(body)
        if not title:
            continue

        path = " > ".join(stack[level] for level in sorted(stack))
        items.append(
            CatalogItem(
                title=title,
                priority=bullet.group("priority"),
                doc_type=bullet.group("doc_type").strip(),
                body=body,
                category=category_from_stack(stack),
                catalog_path=path,
                top_code=top_code,
                line_no=line_no,
            )
        )

    summary = {
        "catalog_items": len(items),
        "skipped_internal_items": skipped_internal,
        "source": CATALOG_PATH.name,
    }
    return items, summary


def load_legacy_rows() -> list[dict[str, Any]]:
    candidates = [
        META_ROOT / "rules_index.json",
    ]
    for path in candidates:
        if path.is_file():
            return json.loads(path.read_text(encoding="utf-8"))
    return []


def legacy_maps(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    by_key: dict[str, dict[str, Any]] = {}
    for row in rows:
        title = row.get("title", "")
        for candidate in title_candidates(title):
            by_key.setdefault(candidate, row)
    return by_key


def infer_tags(text: str, vocabulary: tuple[str, ...]) -> list[str]:
    tags = []
    for tag in vocabulary:
        if tag in text:
            normalized = {
                "保险资管": "保险资金",
                "社保": "养老金社保年金",
                "养老金": "养老金社保年金",
                "年金": "养老金社保年金",
                "账户": "账户保管",
                "保管": "账户保管",
                "估值": "核算估值",
                "核算": "核算估值",
                "清算": "清算交收",
                "交收": "清算交收",
                "披露": "披露报送",
                "报送": "披露报送",
                "数据": "IT数据",
                "合同": "合同法务",
                "法务": "合同法务",
                "风控": "风控内审",
                "内控": "风控内审",
                "审计": "风控内审",
            }.get(tag, tag)
            tags.append(normalized)
    return tags


def infer_issuer(doc_type: str, title: str) -> str:
    text = f"{doc_type} {title}"
    for hint, issuer in ISSUER_BY_HINT.items():
        if hint in text:
            return issuer
    return ""


def merge_catalog_items(items: list[CatalogItem], legacy_by_key: dict[str, dict[str, Any]]) -> list[RuleRecord]:
    records: dict[str, RuleRecord] = {}
    for item in items:
        key = normalize_title(item.title)
        record = records.get(key)
        if not record:
            record = RuleRecord(key=key)
            record.title = item.title
            record.first_line = item.line_no
            record.priority = item.priority
            record.category = item.category
            record.layer = item.doc_type
            record.doc_type = item.doc_type
            record.issuer = infer_issuer(item.doc_type, item.title)
            record.current_status = "待核验" if item.priority == "P0" else "待扩展"
            records[key] = record

        record.priority = stronger_priority(record.priority, item.priority)
        record.is_core = "是" if record.priority == "P0" else "否"
        record.category = record.category or item.category
        record.layer = join_unique([record.layer, item.doc_type])
        record.doc_type = join_unique([record.doc_type, item.doc_type])
        record.catalog_paths.append(item.catalog_path)

        tag_text = " ".join([item.title, item.doc_type, item.body, item.catalog_path])
        record.product_tags.extend(infer_tags(tag_text, TAG_PRODUCTS))
        record.business_line_tags.extend(infer_tags(tag_text, TAG_LINES))
        record.market_tags.extend(infer_tags(tag_text, TAG_MARKETS))
        record.business_tags.extend([item.doc_type, item.top_code])
        if "关于" in item.body or "职责" in item.body or "规则" in item.body:
            record.key_obligations = record.key_obligations or item.body.strip("。")

    for record in records.values():
        legacy_row = None
        for candidate in title_candidates(record.title):
            legacy_row = legacy_by_key.get(candidate)
            if legacy_row:
                break
        if legacy_row:
            record.legacy_row = legacy_row
            record.legacy_id = legacy_row.get("id", "")
            record.title = legacy_row.get("title") or record.title
            record.issuer = legacy_row.get("issuer") or record.issuer
            record.rule_no = legacy_row.get("rule_no", "")
            record.publish_date = legacy_row.get("publish_date", "")
            record.effective_date = legacy_row.get("effective_date", "")
            record.current_status = legacy_row.get("current_status") or record.current_status
            record.source_type = join_unique(split_field(legacy_row.get("source_type", "")))
            record.source_url = legacy_row.get("source_url", "")
            record.business_tags.extend(split_field(legacy_row.get("business_tags", "")))
            if legacy_row.get("notes"):
                auxiliary_p0 = record.priority == "P0" and legacy_row.get("current_status") == "辅助资料"
                record.notes.extend(clean_notes([legacy_row["notes"]], auxiliary_p0=auxiliary_p0))

    ordered = list(records.values())
    ordered.sort(key=lambda rec: rec.first_line)
    for idx, record in enumerate(ordered, start=1):
        record.id = f"TB{idx:03d}"
    return ordered


def reset_generated_dirs() -> None:
    for path in [RAW_ROOT, TEXT_ROOT]:
        ensure_under(path, ROOT)
        if path.exists():
            shutil.rmtree(path)
        path.mkdir(parents=True, exist_ok=True)
    for path in [META_ROOT, ENTRY_ROOT]:
        path.mkdir(parents=True, exist_ok=True)


def snapshot_generated_dirs() -> Path | None:
    cache_root = Path(tempfile.mkdtemp(prefix="custody_kb_cache_"))
    copied = False
    for source in [RAW_ROOT, TEXT_ROOT, META_ROOT]:
        if not source.exists():
            continue
        target = cache_root / source.name
        shutil.copytree(source, target)
        copied = True
    return cache_root if copied else None


def ensure_category_dirs(records: list[RuleRecord]) -> None:
    for record in records:
        (RAW_ROOT / record.category).mkdir(parents=True, exist_ok=True)
        (TEXT_ROOT / record.category).mkdir(parents=True, exist_ok=True)


def prune_empty_dirs(root: Path) -> int:
    removed = 0
    for path in sorted([p for p in root.rglob("*") if p.is_dir()], key=lambda item: len(item.parts), reverse=True):
        try:
            if not any(path.iterdir()):
                path.rmdir()
                removed += 1
        except OSError:
            continue
    return removed


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def source_for_legacy_path(path_value: str) -> Path | None:
    bases = [ROOT]
    if BUILD_CACHE_ROOT:
        bases.insert(0, BUILD_CACHE_ROOT)
    for base in bases:
        candidate = base / path_value
        if candidate.is_file():
            return candidate
    return None


def copy_legacy_files(record: RuleRecord) -> list[dict[str, str]]:
    if not record.legacy_row:
        return []

    raw_dir = RAW_ROOT / record.category / f"{record.id}_{safe_name(record.title)}"
    text_dir = TEXT_ROOT / record.category / f"{record.id}_{safe_name(record.title)}"
    raw_dir.mkdir(parents=True, exist_ok=True)
    text_dir.mkdir(parents=True, exist_ok=True)

    copied: list[dict[str, str]] = []
    raw_paths: list[str] = []
    text_paths: list[str] = []

    for path_value in split_field(record.legacy_row.get("local_path", "")):
        if is_placeholder_path(path_value) or not has_formal_raw_path(path_value):
            continue
        src = source_for_legacy_path(path_value)
        if not src:
            record.errors.append(f"旧原文不存在：{path_value}")
            continue
        dst = raw_dir / src.name
        shutil.copy2(src, dst)
        raw_paths.append(relpath(dst))
        copied.append(
            {
                "legacy_id": record.legacy_id,
                "new_id": record.id,
                "kind": "raw",
                "old_path": path_value,
                "new_path": relpath(dst),
                "sha256": sha256(dst),
            }
        )

    for path_value in split_field(record.legacy_row.get("text_path", "")):
        if is_placeholder_path(path_value):
            continue
        src = source_for_legacy_path(path_value)
        if not src:
            record.errors.append(f"旧文本不存在：{path_value}")
            continue
        dst = text_dir / src.name
        shutil.copy2(src, dst)
        text_paths.append(relpath(dst))
        copied.append(
            {
                "legacy_id": record.legacy_id,
                "new_id": record.id,
                "kind": "text",
                "old_path": path_value,
                "new_path": relpath(dst),
                "sha256": sha256(dst),
            }
        )

    for raw_path in list(raw_paths):
        raw_file = ROOT / raw_path
        if raw_file.suffix.lower() not in [".pdf", ".docx", ".xlsx", ".html", ".htm"]:
            continue
        target = text_dir / f"{raw_file.stem}.txt"
        if text_file_has_meaningful_content(target):
            continue
        regenerated = write_text_for_raw(raw_file, text_dir)
        if regenerated:
            text_paths.append(relpath(regenerated))
            copied.append(
                {
                    "legacy_id": record.legacy_id,
                    "new_id": record.id,
                    "kind": "text-regenerated",
                    "old_path": raw_path,
                    "new_path": relpath(regenerated),
                    "sha256": sha256(regenerated),
                }
            )
            record.notes.append("空文本已从原文重新抽取")

    if raw_paths or text_paths:
        record.notes.append("复用旧公募基金托管库原文/文本")
    record.local_path = join_unique(raw_paths)
    record.text_path = join_unique(text_paths)
    record.file_type = join_unique(sorted({Path(path).suffix.lower().lstrip(".") for path in raw_paths if Path(path).suffix}))
    record.downloaded_count = len(raw_paths)
    return copied


def request_bytes(url: str, *, method: str = "GET", data: bytes | None = None, headers: dict[str, str] | None = None) -> tuple[bytes, dict[str, str]]:
    req_headers = {"User-Agent": USER_AGENT, "Accept": "*/*"}
    if headers:
        req_headers.update(headers)
    current_url = url
    cookie = ""
    for _ in range(3):
        effective_headers = dict(req_headers)
        if cookie:
            effective_headers["Cookie"] = cookie
        req = urllib.request.Request(current_url, data=data, headers=effective_headers, method=method)
        try:
            with urllib.request.urlopen(req, timeout=35) as resp:
                return resp.read(), dict(resp.headers)
        except urllib.error.HTTPError as exc:
            if exc.code not in {301, 302, 303, 307, 308}:
                raise
            set_cookie = exc.headers.get("Set-Cookie", "")
            if set_cookie:
                cookie = set_cookie.split(";", 1)[0]
            location = exc.headers.get("Location")
            if not location:
                raise
            current_url = urllib.parse.urljoin(current_url, location)
        except urllib.error.URLError as exc:
            if isinstance(getattr(exc, "reason", None), ssl.SSLCertVerificationError):
                with urllib.request.urlopen(req, timeout=35, context=ssl._create_unverified_context()) as resp:
                    return resp.read(), dict(resp.headers)
            raise
    raise RuntimeError(f"重定向次数过多：{url}")


def npc_lookup(title: str) -> dict[str, Any] | None:
    payload = {
        "searchRange": 1,
        "sxrq": [],
        "gbrq": [],
        "searchType": 1,
        "sxx": [],
        "gbrqYear": [],
        "flfgCodeId": [],
        "zdjgCodeId": [],
        "searchContent": title,
        "pageNum": 1,
        "pageSize": 10,
    }
    data, _ = request_bytes(
        "https://flk.npc.gov.cn/law-search/search/list",
        method="POST",
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={"Content-Type": "application/json;charset=utf-8", "Accept": "application/json"},
    )
    result = json.loads(data.decode("utf-8"))
    rows = result.get("rows") or []
    plain = lambda s: re.sub(r"<[^>]+>", "", s or "")
    normalized = normalize_title(title)
    for row in rows:
        if normalize_title(plain(row.get("title"))) == normalized:
            return row
    return None


def npc_detail(bbbs: str) -> dict[str, Any] | None:
    url = "https://flk.npc.gov.cn/law-search/search/flfgDetails?" + urllib.parse.urlencode({"bbbs": bbbs})
    data, _ = request_bytes(url, headers={"Accept": "application/json"})
    try:
        return json.loads(data.decode("utf-8")).get("data")
    except json.JSONDecodeError:
        return {
            "_detail_error": "国家法律法规数据库详情接口返回非 JSON，已跳过详情元数据",
            "_response_preview": data[:300].decode("utf-8", errors="ignore"),
        }


def extract_pdf_text(path: Path) -> str:
    try:
        reader = PdfReader(str(path))
        return "\n\n".join((page.extract_text() or "") for page in reader.pages)
    except Exception as exc:
        return f"[PDF文本抽取失败] {exc}"


def extract_docx_text(path: Path) -> str:
    try:
        with zipfile.ZipFile(path) as zf:
            xml = zf.read("word/document.xml")
        doc = etree.fromstring(xml)
        texts = doc.xpath("//*[local-name()='t']/text()")
        return "\n".join(t.strip() for t in texts if t.strip())
    except Exception as exc:
        return f"[DOCX文本抽取失败] {exc}"


def extract_xlsx_text(path: Path) -> str:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"【工作表】{sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value not in (None, "")]
                if values:
                    lines.append("\t".join(values))
            lines.append("")
        workbook.close()
        return "\n".join(lines).strip()
    except Exception as exc:
        return f"[XLSX文本抽取失败] {exc}"


def has_meaningful_text(text: str, min_chars: int = 80) -> bool:
    if (
        text.startswith("[PDF文本抽取失败]")
        or text.startswith("[DOCX文本抽取失败]")
        or text.startswith("[XLSX文本抽取失败]")
    ):
        return False
    return len(re.sub(r"\s+", "", text)) >= min_chars


def text_file_has_meaningful_content(path: Path) -> bool:
    if not path.exists():
        return False
    try:
        return has_meaningful_text(path.read_text(encoding="utf-8", errors="ignore"))
    except OSError:
        return False


FORMAL_RAW_EXTS = {".pdf", ".docx", ".doc", ".xlsx", ".html", ".htm", ".txt"}


def has_formal_raw_path(path_value: str) -> bool:
    return Path(path_value).suffix.lower() in FORMAL_RAW_EXTS


def record_has_formal_raw(record: RuleRecord) -> bool:
    return any(has_formal_raw_path(path) for path in split_field(record.local_path))


def is_placeholder_path(path_value: str) -> bool:
    path = Path(path_value)
    stem = path.stem
    return path.suffix.lower() == ".json" or stem.endswith("_官方入口") or stem.endswith("_npc_detail")


def write_text_for_raw(raw_file: Path, text_dir: Path) -> Path | None:
    if raw_file.suffix.lower() == ".pdf":
        text = extract_pdf_text(raw_file)
    elif raw_file.suffix.lower() == ".docx":
        text = extract_docx_text(raw_file)
    elif raw_file.suffix.lower() == ".xlsx":
        text = extract_xlsx_text(raw_file)
    elif raw_file.suffix.lower() in [".html", ".htm"]:
        text = html_text_from_bytes(raw_file.read_bytes())
    else:
        return None
    if not has_meaningful_text(text):
        return None
    text_dir.mkdir(parents=True, exist_ok=True)
    target = text_dir / f"{raw_file.stem}.txt"
    target.write_text(text, encoding="utf-8", newline="\n")
    return target


def is_composite_title(title: str) -> bool:
    if any(keyword in title for keyword in COMPOSITE_TITLE_KEYWORDS):
        return True
    if re.search(r"[、/]+", title):
        return True
    if re.search(r"(等|及).*(规则|办法|通知|指引|细则|文件)", title):
        return True
    return False


def official_entry_urls(record: RuleRecord) -> list[str]:
    text = f"{record.doc_type} {record.title}"
    urls: list[str] = []
    for keywords, candidates in OFFICIAL_ENTRY_URLS:
        if any(keyword in text for keyword in keywords):
            urls.extend(candidates)
    category_text = f"{record.category} {' '.join(record.catalog_paths)}"
    if any(word in record.doc_type for word in ["法律", "行政法规", "司法解释"]):
        urls.append("https://flk.npc.gov.cn/")
    if "证监会" in record.doc_type or "基金" in text or "证券投资基金" in text or "公募" in text:
        urls.append("https://www.csrc.gov.cn/")
    if "自律规则" in record.doc_type or "私募" in text or "中基协" in text:
        urls.append("https://www.amac.org.cn/")
    if "交易所" in text or "交易所" in category_text:
        urls.extend(["https://www.sse.com.cn/", "https://www.szse.cn/", "https://www.bse.cn/"])
    if "中国结算" in text or "中登" in text or "中国证券登记结算" in text or "中国证券登记结算" in category_text:
        urls.append("https://www.chinaclear.cn/")
    if "中央结算" in text or "中债" in text or "中央结算" in category_text:
        urls.append("https://www.chinabond.com.cn/")
    if "上海清算" in text or "上清所" in text or "上海清算" in category_text:
        urls.append("https://www.shclearing.com.cn/")
    if "交易商协会" in record.doc_type or "交易商协会" in category_text:
        urls.append("https://www.nafmii.org.cn/")
    if "外汇" in text or "QDII" in text or "QFII" in text or "跨境" in category_text:
        urls.append("https://www.safe.gov.cn/")
    if "税" in text or "税务" in record.doc_type or "税务" in category_text:
        urls.extend(["https://www.chinatax.gov.cn/", "https://www.mof.gov.cn/"])
    if "银行" in text or "理财" in text or "信托" in text or "保险" in text:
        urls.append("https://www.nfra.gov.cn/")
    return urls


def should_seed_official_entry(record: RuleRecord) -> bool:
    if record_has_formal_raw(record) or not official_entry_urls(record):
        return False
    text = f"{record.doc_type} {record.title}"
    if any(word in text for word in ["入口", "官方库", "栏目", "平台"]):
        return True
    return True


def seed_official_entry(record: RuleRecord) -> bool:
    urls = official_entry_urls(record)
    if not urls:
        return False
    if record.errors:
        record.notes.extend([f"自动下载未完成，保留官方入口线索：{error}" for error in record.errors])
        record.errors.clear()
    record.source_url = join_unique(split_field(record.source_url) + urls)
    record.source_type = join_unique(split_field(record.source_type) + ["官方入口"])
    record.notes = clean_notes(record.notes, auxiliary_p0=record.priority == "P0")
    record.notes.append("官方入口仅记入台账；不作为本地正式原文")
    return True


def load_source_overrides() -> list[dict[str, Any]]:
    if not SOURCE_OVERRIDES_PATH.exists():
        return []
    return json.loads(SOURCE_OVERRIDES_PATH.read_text(encoding="utf-8"))


def source_override_for_record(record: RuleRecord, overrides: list[dict[str, Any]]) -> dict[str, Any] | None:
    record_title = normalize_title(record.title)
    for item in overrides:
        if item.get("id") and item["id"] == record.id:
            return item
        if item.get("title") and normalize_title(item["title"]) == record_title:
            return item
    return None


def header_value(headers: dict[str, str], name: str) -> str:
    name = name.lower()
    for key, value in headers.items():
        if key.lower() == name:
            return value
    return ""


def extension_from_content_disposition(headers: dict[str, str]) -> str:
    disposition = header_value(headers, "Content-Disposition")
    if not disposition:
        return ""
    match = re.search(r"filename\*?=(?:UTF-8''|\"?)([^\";]+)", disposition, re.I)
    if not match:
        return ""
    filename = urllib.parse.unquote(match.group(1).strip().strip('"'))
    suffix = Path(filename).suffix.lower()
    return suffix if suffix in [".pdf", ".docx", ".doc", ".xlsx", ".html", ".htm", ".json", ".txt"] else ""


def office_zip_extension(data: bytes) -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            names = set(zf.namelist())
        if any(name.startswith("xl/") for name in names):
            return ".xlsx"
        if "word/document.xml" in names:
            return ".docx"
    except Exception:
        return ""
    return ""


def extension_from_payload(data: bytes, headers: dict[str, str], url: str) -> str:
    content_type = header_value(headers, "Content-Type").lower()
    suffix = Path(urllib.parse.unquote(urllib.parse.urlparse(url).path)).suffix.lower()
    disposition_suffix = extension_from_content_disposition(headers)
    if data.startswith(b"%PDF") or "application/pdf" in content_type:
        return ".pdf"
    if disposition_suffix:
        return ".html" if disposition_suffix == ".htm" else disposition_suffix
    if data.startswith(b"PK"):
        return office_zip_extension(data) or ".docx"
    if "spreadsheetml.sheet" in content_type:
        return ".xlsx"
    if "wordprocessingml.document" in content_type:
        return ".docx"
    if data.startswith(b"\xd0\xcf\x11\xe0") or "msword" in content_type:
        return ".doc"
    if "json" in content_type:
        return ".json"
    if "html" in content_type or suffix in [".html", ".htm"] or re.search(br"<html|<!doctype html", data[:500], re.I):
        return ".html"
    if suffix in [".pdf", ".docx", ".doc", ".xlsx", ".html", ".htm", ".json", ".txt"]:
        return ".html" if suffix == ".htm" else suffix
    return ".html"


def decode_html_bytes(data: bytes) -> str:
    charset_match = re.search(br"charset=([A-Za-z0-9_\-]+)", data[:1000], re.I)
    candidates: list[str] = []
    if charset_match:
        declared = charset_match.group(1).decode("ascii", errors="ignore").lower()
        candidates.append({"gb_2312-80": "gb2312"}.get(declared, declared))
    candidates.extend(["utf-8", "gb18030", "gbk", "gb2312"])
    seen: set[str] = set()
    for encoding in candidates:
        if not encoding or encoding in seen:
            continue
        seen.add(encoding)
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def html_text_from_bytes(data: bytes) -> str:
    doc = lxml_html.fromstring(decode_html_bytes(data))
    for node in doc.xpath("//script|//style|//noscript"):
        node.drop_tree()
    lines = [html.unescape(line.strip()) for line in doc.text_content().splitlines()]
    cleaned: list[str] = []
    previous_blank = False
    for line in lines:
        line = re.sub(r"\s+", " ", line)
        if not line:
            if not previous_blank:
                cleaned.append("")
            previous_blank = True
            continue
        cleaned.append(line)
        previous_blank = False
    return "\n".join(cleaned).strip()


def attachment_links_from_html(data: bytes, base_url: str, limit: int = 6) -> list[str]:
    try:
        doc = lxml_html.fromstring(data)
    except Exception:
        return []
    links: list[str] = []
    for anchor in doc.xpath("//a[@href]"):
        href = (anchor.get("href") or "").strip()
        if not href or href.startswith(("javascript:", "#", "mailto:")):
            continue
        label = re.sub(r"\s+", " ", anchor.text_content() or "").strip()
        full = urllib.parse.urljoin(base_url, href)
        marker = f"{full} {label}".lower()
        looks_like_attachment = (
            re.search(r"\.(pdf|docx?|xlsx?|wps)(\?|#|$)", marker)
            or "附件" in label
            or "下载" in label
            or "download" in marker
            or "p020" in marker
        )
        if looks_like_attachment and full not in links:
            links.append(full)
        if len(links) >= limit:
            break
    return links


def write_source_payload(data: bytes, headers: dict[str, str], url: str, raw_file: Path, text_dir: Path) -> tuple[str | None, str | None, list[str]]:
    ext = extension_from_payload(data, headers, url)
    if ext == ".json":
        return None, None, [f"{url} 返回 JSON，未作为法规原文保存"]

    raw_file = raw_file.with_suffix(ext)
    raw_file.write_bytes(data)

    notes: list[str] = []
    text_path: Path | None = None
    if ext in [".pdf", ".docx", ".xlsx"]:
        text_path = write_text_for_raw(raw_file, text_dir)
        if not text_path:
            notes.append(f"{raw_file.name} 未抽取到有效文本")
    elif ext == ".html":
        try:
            text = html_text_from_bytes(data)
            if has_meaningful_text(text):
                text_dir.mkdir(parents=True, exist_ok=True)
                text_path = text_dir / f"{raw_file.stem}.txt"
                text_path.write_text(text, encoding="utf-8", newline="\n")
            else:
                notes.append(f"{raw_file.name} HTML正文过短，未作为可检索文本")
        except Exception as exc:
            notes.append(f"{raw_file.name} HTML文本抽取失败：{exc}")
    else:
        notes.append(f"{raw_file.name} 文件类型 {ext} 已保存但暂未抽取文本")

    return relpath(raw_file), relpath(text_path) if text_path else None, notes


def download_source_url(record: RuleRecord, url: str, raw_dir: Path, text_dir: Path, index: int) -> tuple[list[str], list[str], list[str]]:
    raw_paths: list[str] = []
    text_paths: list[str] = []
    notes: list[str] = []
    data, headers = request_bytes(
        url,
        headers={
            "Accept": "text/html,application/xhtml+xml,application/pdf,application/vnd.openxmlformats-officedocument.wordprocessingml.document,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
        },
    )
    raw_file = raw_dir / f"{record.id}_{safe_name(record.title)}_source{index:02d}"
    raw_path, text_path, payload_notes = write_source_payload(data, headers, url, raw_file, text_dir)
    if raw_path:
        raw_paths.append(raw_path)
    if text_path:
        text_paths.append(text_path)
    notes.extend(payload_notes)

    if extension_from_payload(data, headers, url) == ".html":
        for attachment_index, attachment_url in enumerate(attachment_links_from_html(data, url), start=1):
            try:
                att_data, att_headers = request_bytes(attachment_url, headers={"Accept": "application/pdf,application/vnd.openxmlformats-officedocument.wordprocessingml.document,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*"})
                att_file = raw_dir / f"{record.id}_{safe_name(record.title)}_source{index:02d}_attachment{attachment_index:02d}"
                att_raw, att_text, att_notes = write_source_payload(att_data, att_headers, attachment_url, att_file, text_dir)
                raw_paths.append(att_raw)
                if att_text:
                    text_paths.append(att_text)
                notes.extend(att_notes)
            except Exception as exc:
                notes.append(f"附件下载失败：{attachment_url}；{exc}")
            time.sleep(0.2)

    return raw_paths, text_paths, notes


def apply_source_override(record: RuleRecord, overrides: list[dict[str, Any]]) -> bool:
    item = source_override_for_record(record, overrides)
    if not item:
        return False
    urls = [url for url in item.get("urls", []) if url]
    reference_urls = [url for url in item.get("reference_urls", []) if url]
    if not urls:
        record.errors.append("source_overrides 未配置 urls")
        return False
    had_formal_raw = record_has_formal_raw(record)

    def refresh_override_metadata(success_urls: list[str]) -> None:
        record.source_url = join_unique(split_field(record.source_url) + reference_urls + success_urls)
        record.source_type = join_unique(split_field(record.source_type) + [item.get("source_type", "官方来源")])
        record.current_status = item.get("current_status") or "现行有效"
        if record.current_status not in ["待核验", "待扩展"]:
            stale_fragments = [
                "自动下载未完成，已转官方入口辅助资料",
                "组合型P0目录已入官方入口",
                "现行适用关系待核验",
                "衔接关系待核验",
                "尚无本地正式原文",
            ]
            record.notes = [
                note
                for note in record.notes
                if not any(fragment in note for fragment in stale_fragments)
            ]
        override_notes = item.get("notes", [])
        if isinstance(override_notes, str):
            override_notes = [override_notes]
        record.notes.extend([note for note in override_notes if note])
        record.notes.append("按 source_overrides 官方来源入库")

    if had_formal_raw:
        refresh_override_metadata(urls)
        return True

    raw_dir = RAW_ROOT / record.category / f"{record.id}_{safe_name(record.title)}"
    text_dir = TEXT_ROOT / record.category / f"{record.id}_{safe_name(record.title)}"
    raw_dir.mkdir(parents=True, exist_ok=True)
    text_dir.mkdir(parents=True, exist_ok=True)

    raw_paths: list[str] = []
    text_paths: list[str] = []
    success_urls: list[str] = []
    failures: list[str] = []
    for index, url in enumerate(urls, start=1):
        try:
            new_raw, new_text, notes = download_source_url(record, url, raw_dir, text_dir, index)
            raw_paths.extend(new_raw)
            text_paths.extend(new_text)
            if new_raw:
                success_urls.append(url)
            record.notes.extend(notes)
        except Exception as exc:
            failures.append(f"官方来源下载失败：{url}；{exc}")
        time.sleep(0.2)

    if not raw_paths:
        if had_formal_raw:
            raw_paths = []
            success_urls = urls
        else:
            record.errors.extend(failures)
            return False
    if failures:
        record.notes.extend(failures)

    record.source_url = join_unique(split_field(record.source_url) + success_urls)
    record.source_type = join_unique(split_field(record.source_type) + [item.get("source_type", "官方来源")])
    record.local_path = join_unique(split_field(record.local_path) + raw_paths)
    record.text_path = join_unique(split_field(record.text_path) + text_paths)
    record.file_type = join_unique(sorted({Path(path).suffix.lower().lstrip(".") for path in split_field(record.local_path)}))
    record.downloaded_count = len(split_field(record.local_path))
    refresh_override_metadata(success_urls)
    if not text_paths:
        record.errors.append("官方来源已保存原文但未抽取到可检索文本")
    return True


def derived_source_matches(record: RuleRecord, records: list[RuleRecord]) -> list[RuleRecord]:
    title = record.title
    if not any(marker in title for marker in ["关于", "合同编", "及", "、"]):
        return []
    record_norm = normalize_title(title)
    matches: list[RuleRecord] = []
    for base in records:
        if base is record or not base.local_path:
            continue
        if base.current_status in ["待核验", "待扩展", "历史失效", "辅助资料"]:
            continue
        base_norm = normalize_title(base.title)
        if len(base_norm) < 5 and not record_norm.startswith(base_norm):
            continue
        if record_norm.startswith(base_norm) or base_norm in record_norm:
            matches.append(base)
    matches.sort(key=lambda item: len(normalize_title(item.title)), reverse=True)
    return matches[:4]


def inherit_base_sources(records: list[RuleRecord]) -> None:
    for record in records:
        if record.local_path:
            continue
        matches = derived_source_matches(record, records)
        if not matches:
            continue
        record.source_url = join_unique(split_field(record.source_url) + [url for match in matches for url in split_field(match.source_url)])
        record.source_type = join_unique(split_field(record.source_type) + [f"引用基础法规：{match.id}" for match in matches])
        record.local_path = join_unique([path for match in matches for path in split_field(match.local_path)])
        record.text_path = join_unique([path for match in matches for path in split_field(match.text_path)])
        record.file_type = join_unique(sorted({part for match in matches for part in split_field(match.file_type)}))
        record.downloaded_count = sum(len(split_field(match.local_path)) for match in matches)
        record.current_status = "现行有效"
        record.notes.append("专题拆分项引用已入库基础法规原文：" + "、".join(f"{match.id} {match.title}" for match in matches))
        record.errors = [error for error in record.errors if not error.startswith("国家法律法规数据库未检索到")]


def is_single_normative_record(record: RuleRecord) -> bool:
    text = f"{record.doc_type} {record.title}"
    if any(word in text for word in ["入口", "官方库", "栏目", "平台"]):
        return False
    if is_composite_title(record.title):
        return False
    if any(word in text for word in ["行业手册", "操作实务手册", "辅助资料", "政策解读", "起草说明", "答记者问"]):
        return False
    return any(
        word in text
        for word in [
            "法律",
            "行政法规",
            "司法解释",
            "监管规则",
            "部门规章",
            "管理办法",
            "规定",
            "通知",
            "指引",
            "准则",
            "细则",
        ]
    )


def should_try_npc(record: RuleRecord) -> bool:
    if record_has_formal_raw(record):
        return False
    if not is_single_normative_record(record):
        return False
    return True


def download_npc(record: RuleRecord) -> None:
    title = canonical_title(record.title)
    raw_dir = RAW_ROOT / record.category / f"{record.id}_{safe_name(record.title)}"
    text_dir = TEXT_ROOT / record.category / f"{record.id}_{safe_name(record.title)}"
    raw_dir.mkdir(parents=True, exist_ok=True)
    text_dir.mkdir(parents=True, exist_ok=True)
    try:
        row = npc_lookup(title)
        if not row:
            record.errors.append(f"国家法律法规数据库未检索到：{title}")
            return
        bbbs = row["bbbs"]
        npc_detail(bbbs)
        raw_paths: list[str] = []
        text_paths: list[str] = []
        for fmt in ["pdf", "docx"]:
            try:
                url = f"https://flk.npc.gov.cn/law-search/download/mobile?format={fmt}&bbbs={bbbs}&fileId="
                data, _ = request_bytes(url)
                if fmt == "pdf" and not data.startswith(b"%PDF"):
                    record.notes.append(f"国家法律法规数据库{fmt}下载返回非文件内容，已跳过")
                    continue
                if fmt == "docx" and not data.startswith(b"PK"):
                    record.notes.append(f"国家法律法规数据库{fmt}下载返回非文件内容，已跳过")
                    continue
                target = raw_dir / f"{record.id}_{safe_name(record.title)}_国家法律法规数据库.{fmt}"
                target.write_bytes(data)
                raw_paths.append(relpath(target))
                text_target = write_text_for_raw(target, text_dir)
                if text_target:
                    text_paths.append(relpath(text_target))
            except Exception as exc:
                record.notes.append(f"国家法律法规数据库{fmt}下载失败：{exc}")
            time.sleep(0.2)
        if not raw_paths:
            record.errors.append(f"国家法律法规数据库未取得可用正文文件：{title}")
            return
        record.source_url = join_unique([record.source_url, "https://flk.npc.gov.cn/"])
        record.source_type = join_unique(split_field(record.source_type) + ["国家法律法规数据库"])
        record.local_path = join_unique(split_field(record.local_path) + raw_paths)
        record.text_path = join_unique(split_field(record.text_path) + text_paths)
        record.file_type = join_unique(sorted({Path(path).suffix.lower().lstrip(".") for path in split_field(record.local_path)}))
        record.downloaded_count = len(split_field(record.local_path))
        record.current_status = "现行有效"
        record.notes.append("从国家法律法规数据库下载")
    except Exception as exc:
        record.errors.append(f"国家法律法规数据库下载失败：{exc}")


def record_to_row(record: RuleRecord) -> dict[str, Any]:
    apply_classification_guardrail(record)
    normalize_important_reference_handbook(record)
    clean_stale_auxiliary_notes(record)
    record.business_tags = [tag for tag in record.business_tags if tag != "辅助资料"]
    if (
        not record.local_path
        and record.current_status != "历史失效"
        and record.record_role not in ["官方入口", "重要参考规则", "规则组索引"]
    ):
        record.notes = [note for note in record.notes if note != "复用旧公募基金托管库原文/文本"]
        note = "尚无本地正式原文，待月度巡检"
        if note not in record.notes:
            record.notes.append(note)
    if (
        record.priority == "P0"
        and not record.source_url
        and record.current_status not in ["历史失效", "不适用"]
        and record.record_role not in ["官方入口", "重要参考规则", "规则组索引"]
    ):
        record.current_status = "待核验"
    if "flk.npc.gov.cn" in record.source_url and not record.source_type:
        record.source_type = "国家法律法规数据库"

    return {
        "id": record.id,
        "legacy_id": record.legacy_id,
        "priority": record.priority,
        "category": record.category,
        "title": record.title,
        "layer": record.layer,
        "doc_type": record.doc_type,
        "issuer": record.issuer,
        "rule_no": record.rule_no,
        "publish_date": record.publish_date,
        "effective_date": record.effective_date,
        "current_status": record.current_status,
        "record_role": record.record_role,
        "ingest_status": record.ingest_status,
        "is_core": "是" if record.priority == "P0" else "否",
        "source_type": record.source_type,
        "product_tags": join_unique(record.product_tags),
        "business_line_tags": join_unique(record.business_line_tags),
        "market_tags": join_unique(record.market_tags),
        "business_tags": join_unique(record.business_tags + record.product_tags + record.business_line_tags + record.market_tags),
        "catalog_paths": join_unique(record.catalog_paths),
        "key_obligations": record.key_obligations,
        "source_url": record.source_url,
        "local_path": record.local_path,
        "text_path": record.text_path,
        "file_type": record.file_type,
        "downloaded_count": record.downloaded_count,
        "notes": join_unique(record.notes),
        "errors": join_unique(record.errors),
    }


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def html_anchor(value: str) -> str:
    return re.sub(r"[^0-9A-Za-z\u4e00-\u9fff_-]+", "-", value).strip("-") or "item"


def local_file_href(path: str) -> str:
    href = "../" + path.replace("\\", "/")
    return urllib.parse.quote(href, safe="/._-~%()（）《》【】[];：:—+")


def write_html_index(rows: list[dict[str, Any]], now: str) -> None:
    def unique(values: list[str]) -> list[str]:
        seen: set[str] = set()
        result: list[str] = []
        for value in values:
            if value and value not in seen:
                seen.add(value)
                result.append(value)
        return result

    def file_payload(path: str) -> dict[str, str]:
        suffix = Path(path).suffix.lower().lstrip(".") or "file"
        return {
            "name": Path(path).name,
            "path": path,
            "href": local_file_href(path),
            "type": suffix.upper(),
        }

    browser_rows: list[dict[str, Any]] = []
    for row in rows:
        source_urls = [url for url in unique(split_field(row.get("source_url", ""))) if url.startswith("http")]
        browser_rows.append(
            {
                "id": row.get("id", ""),
                "title": row.get("title", ""),
                "priority": row.get("priority", ""),
                "status": row.get("current_status", ""),
                "role": row.get("record_role", "") or "未分类",
                "ingest": row.get("ingest_status", "") or "未标注",
                "category": row.get("category", ""),
                "catalogs": row.get("catalog_paths", ""),
                "docType": row.get("doc_type", ""),
                "issuer": row.get("issuer", ""),
                "ruleNo": row.get("rule_no", ""),
                "publishDate": row.get("publish_date", ""),
                "effectiveDate": row.get("effective_date", ""),
                "productTags": row.get("product_tags", ""),
                "lineTags": row.get("business_line_tags", ""),
                "marketTags": row.get("market_tags", ""),
                "businessTags": row.get("business_tags", ""),
                "notes": row.get("notes", ""),
                "errors": row.get("errors", ""),
                "sources": source_urls,
                "rawFiles": [file_payload(path) for path in split_field(row.get("local_path", ""))],
                "textFiles": [file_payload(path) for path in split_field(row.get("text_path", ""))],
            }
        )

    data_json = json.dumps(browser_rows, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")
    metrics = [
        ("总收录记录", len(rows)),
        ("正式规则", sum(1 for row in rows if row.get("record_role") == "正式规则")),
        ("重要参考", sum(1 for row in rows if row.get("record_role") == "重要参考规则")),
        ("规则组索引", sum(1 for row in rows if row.get("record_role") == "规则组索引")),
        ("有本地原文文件", sum(1 for row in rows if row.get("local_path"))),
        ("官方来源覆盖", sum(1 for row in rows if row.get("source_url"))),
        (
            "待处理",
            sum(
                1
                for row in rows
                if row["current_status"] in ["待核验", "待扩展"]
                or str(row.get("ingest_status", "")).startswith("待")
            ),
        ),
    ]
    metrics_html = "".join(
        f'<div class="metric"><strong>{count}</strong><span>{html.escape(label)}</span></div>'
        for label, count in metrics
    )

    html_doc = """<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>商业银行托管业务法规库总目录</title>
  <style>
    :root {
      --paper:#f7f3ea;
      --panel:#fffdf8;
      --ink:#24221e;
      --muted:#746f66;
      --line:#d8d0c1;
      --line-strong:#b9ac96;
      --teal:#155f67;
      --red:#9f3d32;
      --gold:#b17a25;
      --green:#35745d;
      --shadow:0 16px 36px rgba(64,50,31,.10);
    }
    * { box-sizing:border-box; }
    body {
      margin:0;
      color:var(--ink);
      background:
        linear-gradient(90deg, rgba(36,34,30,.045) 1px, transparent 1px),
        linear-gradient(180deg, rgba(36,34,30,.035) 1px, transparent 1px),
        var(--paper);
      background-size:28px 28px;
      font-family:"Microsoft YaHei UI","Microsoft YaHei","PingFang SC",sans-serif;
      font-size:14px;
    }
    a { color:var(--teal); text-decoration:none; }
    a:hover { text-decoration:underline; }
    button, input, select { font:inherit; }
    .shell {
      display:grid;
      grid-template-columns:312px minmax(0,1fr);
      min-height:100vh;
    }
    .rail {
      position:sticky;
      top:0;
      height:100vh;
      overflow:auto;
      border-right:1px solid var(--line);
      background:rgba(255,253,248,.86);
      backdrop-filter:blur(12px);
      padding:18px;
    }
    .brand {
      border-bottom:2px solid var(--ink);
      padding-bottom:14px;
      margin-bottom:14px;
    }
    .brand h1 {
      margin:0;
      font-family:"Source Han Serif SC","Noto Serif CJK SC","SimSun",serif;
      font-size:25px;
      line-height:1.15;
      font-weight:700;
    }
    .brand p { margin:8px 0 0; color:var(--muted); font-size:12px; }
    .metrics {
      display:grid;
      grid-template-columns:repeat(2,minmax(0,1fr));
      gap:8px;
      margin-bottom:18px;
    }
    .metric {
      min-height:58px;
      border:1px solid var(--line);
      background:var(--panel);
      padding:10px;
      box-shadow:0 1px 0 rgba(255,255,255,.8) inset;
    }
    .metric strong { display:block; font-size:22px; line-height:1; color:var(--teal); }
    .metric span { display:block; margin-top:6px; font-size:12px; color:var(--muted); }
    .tree-title {
      display:flex;
      align-items:center;
      justify-content:space-between;
      gap:12px;
      margin:14px 0 8px;
      color:var(--muted);
      font-size:12px;
    }
    .tree-all,
    .tree button,
    .tree summary {
      width:100%;
      min-height:32px;
      border:0;
      border-left:3px solid transparent;
      background:transparent;
      color:var(--ink);
      cursor:pointer;
      display:flex;
      align-items:center;
      justify-content:space-between;
      gap:10px;
      padding:6px 8px;
      text-align:left;
      list-style:none;
    }
    .tree summary::-webkit-details-marker { display:none; }
    .tree button { padding-left:calc(10px + var(--depth,0) * 12px); }
    .tree summary { padding-left:calc(8px + var(--depth,0) * 12px); }
    .tree span { overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
    .tree b {
      flex:0 0 auto;
      min-width:28px;
      color:var(--muted);
      font-size:12px;
      font-weight:600;
      text-align:right;
    }
    .tree [data-category].is-active {
      border-left-color:var(--red);
      background:#f2e6d2;
      color:#111;
    }
    .tree details { margin:1px 0; }
    .tree details > div { margin-left:4px; }
    .main {
      min-width:0;
      padding:22px min(3vw,34px) 34px;
    }
    .mast {
      display:flex;
      align-items:flex-end;
      justify-content:space-between;
      gap:24px;
      padding-bottom:16px;
      border-bottom:2px solid var(--ink);
    }
    .mast h2 {
      margin:0;
      font-family:"Source Han Serif SC","Noto Serif CJK SC","SimSun",serif;
      font-size:31px;
      line-height:1.12;
      font-weight:700;
    }
    .mast p { margin:7px 0 0; color:var(--muted); }
    .stamp {
      flex:0 0 auto;
      border:1px solid var(--line-strong);
      color:var(--red);
      padding:8px 10px;
      font-size:12px;
      background:rgba(255,253,248,.72);
    }
    .toolbar {
      display:grid;
      grid-template-columns:minmax(220px,1fr) repeat(4,minmax(112px,148px)) auto;
      gap:10px;
      align-items:center;
      margin:16px 0;
    }
    .toolbar input,
    .toolbar select,
    .toolbar button {
      height:38px;
      border:1px solid var(--line-strong);
      background:var(--panel);
      color:var(--ink);
      border-radius:0;
      padding:0 10px;
    }
    .toolbar button {
      cursor:pointer;
      background:#efe4d0;
      color:#2d2416;
    }
    .workspace {
      display:grid;
      grid-template-columns:minmax(0,1fr) 360px;
      gap:14px;
      align-items:start;
    }
    .list-panel,
    .detail {
      border:1px solid var(--line);
      background:rgba(255,253,248,.94);
      box-shadow:var(--shadow);
    }
    .list-head {
      display:flex;
      justify-content:space-between;
      align-items:center;
      gap:14px;
      min-height:48px;
      border-bottom:1px solid var(--line);
      padding:10px 12px;
    }
    .list-head strong { color:var(--teal); }
    .list-head span { color:var(--muted); font-size:12px; }
    .table-wrap { overflow:auto; max-height:calc(100vh - 190px); }
    table { width:100%; border-collapse:collapse; table-layout:fixed; }
    th {
      position:sticky;
      top:0;
      z-index:1;
      background:#ede3d1;
      color:#4f4638;
      border-bottom:1px solid var(--line-strong);
      font-size:12px;
      font-weight:700;
      text-align:left;
      padding:9px 8px;
    }
    td {
      border-bottom:1px solid #e6dece;
      padding:9px 8px;
      vertical-align:top;
      line-height:1.35;
    }
    tbody tr { cursor:pointer; }
    tbody tr:hover,
    tbody tr.is-selected { background:#f4ead8; }
    .col-id { width:76px; }
    .col-priority { width:70px; }
    .col-role { width:104px; }
    .col-status { width:92px; }
    .col-files { width:76px; text-align:right; }
    .rule-title { font-weight:700; color:#1f211d; }
    .subline { display:block; margin-top:3px; color:var(--muted); font-size:12px; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
    .chip {
      display:inline-flex;
      align-items:center;
      min-height:22px;
      border:1px solid var(--line-strong);
      padding:2px 7px;
      font-size:12px;
      color:#463d30;
      background:#fffaf0;
      white-space:nowrap;
    }
    .chip.formal { border-color:#87a897; color:var(--green); }
    .chip.reference { border-color:#b79655; color:var(--gold); }
    .chip.index { border-color:#8fb0bd; color:var(--teal); }
    .chip.entry { border-color:#c99892; color:var(--red); }
    .detail {
      position:sticky;
      top:18px;
      max-height:calc(100vh - 36px);
      overflow:auto;
    }
    .detail-head {
      padding:15px;
      border-bottom:1px solid var(--line);
      background:#f1e7d4;
    }
    .detail-head h3 { margin:8px 0 0; font-size:20px; line-height:1.35; }
    .detail-id { font-weight:800; color:var(--red); }
    .detail-body { padding:14px 15px 18px; }
    .kv {
      display:grid;
      grid-template-columns:74px minmax(0,1fr);
      gap:7px 12px;
      margin:0 0 14px;
    }
    .kv dt { color:var(--muted); font-size:12px; }
    .kv dd { margin:0; overflow-wrap:anywhere; }
    .detail h4 {
      margin:18px 0 8px;
      font-size:13px;
      color:#40382c;
      border-bottom:1px solid var(--line);
      padding-bottom:6px;
    }
    .file-list { display:grid; gap:7px; }
    .file-list a {
      display:grid;
      grid-template-columns:44px minmax(0,1fr);
      gap:8px;
      align-items:center;
      border:1px solid var(--line);
      background:#fffaf0;
      padding:7px 8px;
      min-height:38px;
    }
    .file-list b {
      color:#fff;
      background:var(--teal);
      padding:3px 4px;
      text-align:center;
      font-size:11px;
    }
    .file-list span { overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
    .empty { color:#8b6b32; font-size:13px; }
    .note {
      white-space:pre-wrap;
      overflow-wrap:anywhere;
      color:#4b463e;
      line-height:1.55;
    }
    @media (max-width:1180px) {
      .shell { grid-template-columns:260px minmax(0,1fr); }
      .workspace { grid-template-columns:1fr; }
      .detail { position:relative; top:auto; max-height:none; }
    }
    @media (max-width:820px) {
      .shell { display:block; }
      .rail { position:relative; height:auto; max-height:48vh; }
      .main { padding:16px; }
      .mast { display:block; }
      .stamp { display:inline-block; margin-top:12px; }
      .toolbar { grid-template-columns:1fr 1fr; }
      .toolbar input { grid-column:1 / -1; }
      .toolbar button { grid-column:1 / -1; }
      .table-wrap { max-height:none; }
      .col-status, .col-role { width:86px; }
    }
  </style>
</head>
<body>
  <div class="shell">
    <aside class="rail">
      <div class="brand">
        <h1>托管法规库</h1>
        <p>生成时间：__NOW__</p>
      </div>
      <div class="metrics">__METRICS__</div>
      <div class="tree-title"><span>原文库目录</span><span id="activePath">全部</span></div>
      <nav class="tree" id="tree"></nav>
    </aside>
    <main class="main">
      <header class="mast">
        <div>
          <h2>商业银行托管业务法规库总目录</h2>
          <p>按原文库路径浏览，按台账字段筛选，逐条查看原文、文本和官方来源。</p>
        </div>
        <div class="stamp">01 原文库 / 02 文本库 / 03 台账</div>
      </header>
      <section class="toolbar">
        <input id="query" type="search" placeholder="搜索 TB编号、标题、机构、标签、路径或来源">
        <select id="roleFilter"><option value="">全部角色</option></select>
        <select id="statusFilter"><option value="">全部状态</option></select>
        <select id="priorityFilter"><option value="">全部优先级</option></select>
        <select id="fileFilter">
          <option value="">全部原文状态</option>
          <option value="with">有本地原文文件</option>
          <option value="without">无本地原文文件</option>
        </select>
        <button id="reset" type="button">重置</button>
      </section>
      <section class="workspace">
        <div class="list-panel">
          <div class="list-head">
            <div><strong id="resultCount">0</strong> <span>条记录</span></div>
            <span id="resultHint">全部目录</span>
          </div>
          <div class="table-wrap">
            <table>
              <thead>
                <tr>
                  <th class="col-id">ID</th>
                  <th>规则名称</th>
                  <th class="col-priority">优先级</th>
                  <th class="col-role">角色</th>
                  <th class="col-status">状态</th>
                  <th class="col-files">文件</th>
                </tr>
              </thead>
              <tbody id="ruleRows"></tbody>
            </table>
          </div>
        </div>
        <aside class="detail" id="detail"></aside>
      </section>
    </main>
  </div>
  <script id="ruleData" type="application/json">__DATA__</script>
  <script>
    const rules = JSON.parse(document.getElementById('ruleData').textContent);
    const state = { query:'', role:'', status:'', priority:'', file:'', category:'', selectedId:'' };
    const els = {
      tree: document.getElementById('tree'),
      activePath: document.getElementById('activePath'),
      query: document.getElementById('query'),
      role: document.getElementById('roleFilter'),
      status: document.getElementById('statusFilter'),
      priority: document.getElementById('priorityFilter'),
      file: document.getElementById('fileFilter'),
      reset: document.getElementById('reset'),
      rows: document.getElementById('ruleRows'),
      count: document.getElementById('resultCount'),
      hint: document.getElementById('resultHint'),
      detail: document.getElementById('detail')
    };

    function esc(value) {
      return String(value ?? '').replace(/[&<>"']/g, char => ({
        '&':'&amp;', '<':'&lt;', '>':'&gt;', '"':'&quot;', "'":'&#39;'
      }[char]));
    }

    function domain(url) {
      try { return new URL(url).hostname || url; } catch { return url; }
    }

    function roleClass(role) {
      if (role === '正式规则') return 'formal';
      if (role === '重要参考规则') return 'reference';
      if (role === '规则组索引') return 'index';
      if (role === '官方入口') return 'entry';
      return '';
    }

    function fillSelect(select, values) {
      values.filter(Boolean).sort((a, b) => a.localeCompare(b, 'zh-Hans-CN')).forEach(value => {
        const option = document.createElement('option');
        option.value = value;
        option.textContent = value;
        select.appendChild(option);
      });
    }

    function buildTree() {
      const root = { count: rules.length, children: new Map() };
      for (const rule of rules) {
        const parts = rule.category.split('/').filter(Boolean);
        let node = root;
        for (const part of parts) {
          if (!node.children.has(part)) node.children.set(part, { count:0, children:new Map() });
          node = node.children.get(part);
          node.count += 1;
        }
      }
      function renderNode(node, prefix, depth) {
        return [...node.children.entries()]
          .sort((a, b) => a[0].localeCompare(b[0], 'zh-Hans-CN'))
          .map(([name, child]) => {
            const path = [...prefix, name].join('/');
            const label = `<span>${esc(name)}</span><b>${child.count}</b>`;
            if (child.children.size) {
              return `<details ${depth < 1 ? 'open' : ''}><summary data-category="${esc(path)}" style="--depth:${depth}">${label}</summary><div>${renderNode(child, [...prefix, name], depth + 1)}</div></details>`;
            }
            return `<button type="button" data-category="${esc(path)}" style="--depth:${depth}">${label}</button>`;
          }).join('');
      }
      els.tree.innerHTML = `<button type="button" class="tree-all" data-category=""><span>全部目录</span><b>${rules.length}</b></button>${renderNode(root, [], 0)}`;
    }

    function searchable(rule) {
      return [
        rule.id, rule.title, rule.priority, rule.status, rule.role, rule.ingest,
        rule.category, rule.catalogs, rule.docType, rule.issuer, rule.ruleNo,
        rule.productTags, rule.lineTags, rule.marketTags, rule.businessTags,
        rule.notes, rule.sources.join(' '), rule.rawFiles.map(file => file.name).join(' ')
      ].join(' ').toLowerCase();
    }

    for (const rule of rules) rule._search = searchable(rule);

    function filteredRules() {
      const terms = state.query.trim().toLowerCase().split(/\\s+/).filter(Boolean);
      return rules.filter(rule => {
        if (state.category && !rule.category.startsWith(state.category)) return false;
        if (state.role && rule.role !== state.role) return false;
        if (state.status && rule.status !== state.status) return false;
        if (state.priority && rule.priority !== state.priority) return false;
        if (state.file === 'with' && rule.rawFiles.length === 0) return false;
        if (state.file === 'without' && rule.rawFiles.length > 0) return false;
        return terms.every(term => rule._search.includes(term));
      });
    }

    function tagText(rule) {
      return [rule.issuer, rule.ruleNo, rule.docType].filter(Boolean).join(' / ') || rule.category;
    }

    function renderRows(list) {
      els.rows.innerHTML = list.map(rule => {
        const fileCount = rule.rawFiles.length ? `${rule.rawFiles.length} 原文 / ${rule.textFiles.length} 文本` : '无';
        return `<tr data-id="${esc(rule.id)}" class="${rule.id === state.selectedId ? 'is-selected' : ''}">
          <td class="col-id"><strong>${esc(rule.id)}</strong></td>
          <td><span class="rule-title">${esc(rule.title)}</span><span class="subline">${esc(tagText(rule))}</span></td>
          <td class="col-priority">${esc(rule.priority)}</td>
          <td class="col-role"><span class="chip ${roleClass(rule.role)}">${esc(rule.role)}</span></td>
          <td class="col-status">${esc(rule.ingest)}</td>
          <td class="col-files">${esc(fileCount)}</td>
        </tr>`;
      }).join('');
    }

    function fileLinks(files, emptyText) {
      if (!files.length) return `<div class="empty">${esc(emptyText)}</div>`;
      return `<div class="file-list">${files.map(file => `<a href="${esc(file.href)}" target="_blank" rel="noopener" title="${esc(file.path)}"><b>${esc(file.type)}</b><span>${esc(file.name)}</span></a>`).join('')}</div>`;
    }

    function sourceLinks(urls) {
      if (!urls.length) return '<div class="empty">暂无官方来源链接</div>';
      return `<div class="file-list">${urls.map(url => `<a href="${esc(url)}" target="_blank" rel="noopener" title="${esc(url)}"><b>URL</b><span>${esc(domain(url))}</span></a>`).join('')}</div>`;
    }

    function renderDetail(rule) {
      if (!rule) {
        els.detail.innerHTML = '<div class="detail-body empty">当前筛选条件下没有记录</div>';
        return;
      }
      state.selectedId = rule.id;
      els.detail.innerHTML = `<div class="detail-head">
          <div class="detail-id">${esc(rule.id)} · ${esc(rule.priority)}</div>
          <h3>${esc(rule.title)}</h3>
        </div>
        <div class="detail-body">
          <dl class="kv">
            <dt>角色</dt><dd><span class="chip ${roleClass(rule.role)}">${esc(rule.role)}</span></dd>
            <dt>入库</dt><dd>${esc(rule.ingest)}</dd>
            <dt>效力</dt><dd>${esc(rule.status)}</dd>
            <dt>机构</dt><dd>${esc(rule.issuer || '待核验')}</dd>
            <dt>文号</dt><dd>${esc(rule.ruleNo || '待核验')}</dd>
            <dt>类型</dt><dd>${esc(rule.docType || '未标注')}</dd>
            <dt>路径</dt><dd>${esc(rule.category)}</dd>
            <dt>标签</dt><dd>${esc([rule.productTags, rule.lineTags, rule.marketTags, rule.businessTags].filter(Boolean).join(' / ') || '未标注')}</dd>
          </dl>
          <h4>本地原文</h4>${fileLinks(rule.rawFiles, '暂无本地原文')}
          <h4>文本抽取</h4>${fileLinks(rule.textFiles, '暂无文本抽取')}
          <h4>官方来源</h4>${sourceLinks(rule.sources)}
          <h4>目录挂接</h4><div class="note">${esc(rule.catalogs || rule.category)}</div>
          ${rule.notes ? `<h4>备注</h4><div class="note">${esc(rule.notes)}</div>` : ''}
          ${rule.errors ? `<h4>提示</h4><div class="note">${esc(rule.errors)}</div>` : ''}
        </div>`;
      renderRows(filteredRules());
    }

    function updateActiveTree() {
      document.querySelectorAll('[data-category]').forEach(node => {
        node.classList.toggle('is-active', node.dataset.category === state.category);
      });
      els.activePath.textContent = state.category || '全部';
    }

    function render() {
      const list = filteredRules();
      if (!list.some(rule => rule.id === state.selectedId)) {
        state.selectedId = list[0]?.id || '';
      }
      els.count.textContent = list.length;
      els.hint.textContent = state.category || '全部目录';
      renderRows(list);
      renderDetail(list.find(rule => rule.id === state.selectedId));
      updateActiveTree();
    }

    fillSelect(els.role, [...new Set(rules.map(rule => rule.role))]);
    fillSelect(els.status, [...new Set(rules.map(rule => rule.status))]);
    fillSelect(els.priority, [...new Set(rules.map(rule => rule.priority))]);
    buildTree();
    render();

    els.query.addEventListener('input', event => { state.query = event.target.value; render(); });
    els.role.addEventListener('change', event => { state.role = event.target.value; render(); });
    els.status.addEventListener('change', event => { state.status = event.target.value; render(); });
    els.priority.addEventListener('change', event => { state.priority = event.target.value; render(); });
    els.file.addEventListener('change', event => { state.file = event.target.value; render(); });
    els.reset.addEventListener('click', () => {
      state.query = state.role = state.status = state.priority = state.file = state.category = '';
      els.query.value = els.role.value = els.status.value = els.priority.value = els.file.value = '';
      render();
    });
    els.tree.addEventListener('click', event => {
      const node = event.target.closest('[data-category]');
      if (!node) return;
      state.category = node.dataset.category;
      render();
    });
    els.rows.addEventListener('click', event => {
      const row = event.target.closest('tr[data-id]');
      if (!row) return;
      const rule = rules.find(item => item.id === row.dataset.id);
      renderDetail(rule);
    });
  </script>
</body>
</html>
"""
    html_doc = (
        html_doc.replace("__NOW__", html.escape(now))
        .replace("__METRICS__", metrics_html)
        .replace("__DATA__", data_json)
    )
    (ENTRY_ROOT / "总目录.html").write_text(html_doc, encoding="utf-8", newline="\n")


def read_manual_update_log() -> str:
    if not UPDATE_LOG_PATH.exists():
        return DEFAULT_MANUAL_UPDATE_LOG
    text = UPDATE_LOG_PATH.read_text(encoding="utf-8")
    if MANUAL_UPDATE_HEADING not in text:
        return DEFAULT_MANUAL_UPDATE_LOG
    manual = text.split(MANUAL_UPDATE_HEADING, 1)[1]
    if AUTO_UPDATE_HEADING in manual:
        manual = manual.split(AUTO_UPDATE_HEADING, 1)[0]
    manual = manual.strip()
    return manual or DEFAULT_MANUAL_UPDATE_LOG


def write_update_log(rows: list[dict[str, Any]], now: str) -> None:
    manual_log = read_manual_update_log()
    role_counts = {
        role: sum(1 for row in rows if row.get("record_role") == role)
        for role in ["正式规则", "重要参考规则", "规则组索引", "官方入口"]
    }
    waiting_count = sum(
        1
        for row in rows
        if row["current_status"] in ["待核验", "待扩展"]
        or str(row.get("ingest_status", "")).startswith("待")
        or str(row.get("ingest_status", "")).endswith("待复核")
    )
    auto_log = [
        f"- {now}：自动重建商业银行托管业务法规库派生文件，刷新元数据台账、入口索引、待核验清单和搜索索引。",
        (
            f"  - 记录概况：总收录 {len(rows)}；正式规则 {role_counts['正式规则']}；"
            f"重要参考规则 {role_counts['重要参考规则']}；规则组索引 {role_counts['规则组索引']}；"
            f"官方入口 {role_counts['官方入口']}。"
        ),
        (
            f"  - 覆盖情况：有本地原文文件 {sum(1 for row in rows if row.get('local_path'))}；"
            f"官方来源覆盖 {sum(1 for row in rows if row.get('source_url'))}；待处理 {waiting_count}。"
        ),
    ]
    auto_log_text = "\n".join(auto_log)
    UPDATE_LOG_PATH.write_text(
        "# 更新记录\n\n"
        f"{MANUAL_UPDATE_HEADING}\n\n"
        f"{manual_log}\n\n"
        f"{AUTO_UPDATE_HEADING}\n\n"
        f"{auto_log_text}\n",
        encoding="utf-8",
        newline="\n",
    )


def write_markdown_indexes(rows: list[dict[str, Any]]) -> None:
    now = time.strftime("%Y-%m-%d %H:%M:%S")
    table = "\n".join(
        f"| {r['id']} | {r['legacy_id']} | {r['priority']} | {r['category']} | {r['title']} | {r['current_status']} | {r.get('record_role', '')} | {r.get('ingest_status', '')} | {r['product_tags']} | {r['business_line_tags']} | {r['market_tags']} |"
        for r in rows
    )
    (META_ROOT / "rules_index.md").write_text(
        "# 法规元数据台账\n\n"
        f"生成时间：{now}\n\n"
        "| ID | 旧ID | 优先级 | 分类 | 文件名称 | 效力/核验状态 | 记录角色 | 入库状态 | 产品标签 | 条线标签 | 市场标签 |\n"
        "|---|---|---|---|---|---|---|---|---|---|---|\n"
        f"{table}\n",
        encoding="utf-8",
        newline="\n",
    )
    write_html_index(rows, now)
    write_update_log(rows, now)


def write_unresolved(rows: list[dict[str, Any]]) -> None:
    selected = [
        r
        for r in rows
        if r["current_status"] in ["待核验", "待扩展"]
        or r["errors"]
        or str(r.get("ingest_status", "")).startswith("待")
        or str(r.get("ingest_status", "")).endswith("待复核")
    ]
    lines = [
        "# 待核验与未完成项目",
        "",
        "以下项目均标注“待月度巡检”。官方发布项目应按正式规则、重要参考规则、规则组索引或官方入口处理。",
        "",
    ]
    if not selected:
        lines.append("暂无待核验、待扩展或待入库项目。")
    for row in selected:
        role = row.get("record_role", "")
        ingest_status = row.get("ingest_status", "")
        if role == "正式规则":
            source_state = "具约束力规则，需补入官方原文并核验现行性" if not row.get("local_path") else "已入库但需复核效力状态和版本"
        elif role == "规则组索引":
            source_state = "规则集合或规则库索引，需拆分为具体规则后逐条入库"
        elif role == "重要参考规则":
            source_state = "官方发布参考项目，需按元数据或正文保持可检索"
        else:
            source_state = "已联网或本地定位到官方来源但尚未完整入库" if row.get("source_url") else "仅有目录线索，尚待核验"
        lines.append(f"## {row['id']} {row['title']}")
        lines.append(f"- 效力/核验状态：{row['current_status']}")
        lines.append(f"- 记录角色：{role or '未分类'}")
        lines.append(f"- 入库状态：{ingest_status or '未标注'}")
        lines.append(f"- 优先级：{row['priority']}")
        lines.append(f"- 目录挂接：{row['catalog_paths']}")
        lines.append(f"- 缺口类型：{source_state}")
        lines.append("- 巡检标记：待月度巡检")
        lines.append(f"- 来源：{row['source_url'] or '待检索'}")
        if row["notes"]:
            lines.append(f"- 备注：{row['notes']}")
        if row["errors"]:
            lines.append(f"- 错误/提示：{row['errors']}")
        lines.append("")
    UNRESOLVED_PATH.write_text("\n".join(lines), encoding="utf-8", newline="\n")


def main() -> None:
    global BUILD_CACHE_ROOT
    legacy_rows = load_legacy_rows()
    legacy_by_key = legacy_maps(legacy_rows)
    source_overrides = load_source_overrides()
    catalog_items, catalog_summary = parse_catalog()
    records = merge_catalog_items(catalog_items, legacy_by_key)

    BUILD_CACHE_ROOT = snapshot_generated_dirs()
    reset_generated_dirs()
    ensure_category_dirs(records)

    for record in records:
        copy_legacy_files(record)
        apply_source_override(record, source_overrides)
        if should_try_npc(record):
            download_npc(record)
            time.sleep(0.4)
        if should_seed_official_entry(record):
            seed_official_entry(record)

    inherit_base_sources(records)
    rows = [record_to_row(record) for record in records]

    META_ROOT.mkdir(parents=True, exist_ok=True)
    write_csv(META_ROOT / "rules_index.csv", rows, FIELDS)
    (META_ROOT / "rules_index.json").write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8", newline="\n")
    write_markdown_indexes(rows)
    write_unresolved(rows)
    empty_dirs_removed = prune_empty_dirs(RAW_ROOT) + prune_empty_dirs(TEXT_ROOT)
    search_index_summary = build_index(ROOT)

    summary = {
        **catalog_summary,
        "total_rules": len(rows),
        "p0_rules": sum(1 for row in rows if row["priority"] == "P0"),
        "rules_with_files": sum(1 for row in rows if row["local_path"]),
        "rules_with_formal_raw": sum(
            1
            for row in rows
            if any(has_formal_raw_path(path) for path in split_field(row["local_path"]))
        ),
        "json_only_rules": sum(
            1
            for row in rows
            if row["local_path"]
            and not any(has_formal_raw_path(path) for path in split_field(row["local_path"]))
        ),
        "p0_without_files": sum(1 for row in rows if row["priority"] == "P0" and not row["local_path"]),
        "rules_waiting": sum(
            1
            for row in rows
            if row["current_status"] in ["待核验", "待扩展"]
            or str(row.get("ingest_status", "")).startswith("待")
            or str(row.get("ingest_status", "")).endswith("待复核")
        ),
        "record_roles": {
            role: sum(1 for row in rows if row.get("record_role") == role)
            for role in sorted({row.get("record_role", "") for row in rows})
            if role
        },
        "ingest_statuses": {
            status: sum(1 for row in rows if row.get("ingest_status") == status)
            for status in sorted({row.get("ingest_status", "") for row in rows})
            if status
        },
        "legacy_reused_records": sum(1 for row in rows if row["legacy_id"]),
        "empty_dirs_removed": empty_dirs_removed,
        "search_index": search_index_summary,
        "built_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if BUILD_CACHE_ROOT and BUILD_CACHE_ROOT.exists():
        shutil.rmtree(BUILD_CACHE_ROOT, ignore_errors=True)
    BUILD_CACHE_ROOT = None


if __name__ == "__main__":
    main()
